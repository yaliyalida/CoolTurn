from openpyxl import load_workbook
from docx import Document
from docx.oxml.ns import qn
import re
import os


# 转换成word文本
def text(file_path, user_path, data, key):
    # 读取excel xlsx文件
    wb = load_workbook(file_path)

    # 获取所有sheet页名字
    xl_sheet_names = wb.sheetnames

    # 定位到相应sheet页,[0]为sheet页索引
    ws = wb[xl_sheet_names[0]]

    # 获取行列数
    excel_row = ws.max_row

    # 找到模板中需要插入内容的位置
    blanks = re.findall(r'{\d+}', data)

    # 找到key对应的列
    k = 0
    for row in list(ws.rows)[0]:
        k += 1
        if key == row.value:
            break

    # i表示第i行数据
    i = 2

    # 写入word文件
    while i <= excel_row:

        tem = data

        for blank in blanks:
            index = int(blank[1:-1])  # 找到要填入的内容对应excel表格的单元格
            value = str(ws.cell(row=i, column=index).value)  # 找到要替换的内容
            tem = tem.replace(blank, value)  # 替换

        # 文件名
        name = str(ws.cell(row=i, column=k).value)

        # 创建word文档
        document = Document()

        # 设置文档字体
        document.styles['Normal'].font.name = u'宋体'
        document.styles['Normal']._element.rPr.rFonts.set(qn('w:eastAsia'), u'宋体')

        # 向文档中写入内容
        document.add_paragraph(tem)

        # 输出路径的前面一部分到时候用用户输入的
        docx_path = user_path + name + '.docx'
        document.save(docx_path)

        i += 1


# 用户指定要处理的文件路径
file_path = 'C:\\Users\\admin\\Desktop\\学生信息表.xlsx'

# 用户指定处理完成后文件保存路径
user_path = 'C:\\Users\\admin\\Desktop\\'

# 简单的模板，由用户指定
data = "我的名字是{1},学号是{5},性别{3},年龄{2},在{4}学院学习"

# 选定用来命名输出文件的列，也是用户指定
key = '学号'

# 写入word文档
text(file_path, user_path, data, key)
