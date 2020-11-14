from openpyxl import load_workbook
from docx import Document
from docx.oxml.ns import qn


# 转换成word表格
def tables(file_path, user_path, key):
    # 读取excel xlsx文件
    wb = load_workbook(file_path)

    # 获取所有sheet页名字
    xl_sheet_names = wb.sheetnames

    # 定位到相应sheet页,[0]为sheet页索引
    ws = wb[xl_sheet_names[0]]

    # 获取行列数
    excel_row = ws.max_row
    excel_column = ws.max_column

    # 将excel表格每一行转换成列表，方便操作
    rows = list(ws.rows)

    # 找到key对应的列
    k = 0
    for row in rows[0]:
        k += 1
        if key == row.value:
            break

    # 写入word文件
    i = 1
    while i < excel_row:
        # 创建word文档
        document = Document()

        # 在word文档中添加表格
        table = document.add_table(rows=1, cols=excel_column, style="Table Grid")

        # 添加表头
        hdr_cells = table.rows[0].cells
        for num in range(0, excel_column):
            hdr_cells[num].text = u'' + str(rows[0][num].value)

        # 插入每一行
        row_cells = table.add_row().cells
        for num in range(0, excel_column):
            content = str(rows[i][num].value)
            content = "" if content == "None" else content
            row_cells[num].text = u'' + content

        # 文件名
        name = str(ws.cell(row=i + 1, column=k).value)

        # 设置文档字体
        document.styles['Normal'].font.name = u'宋体'
        document.styles['Normal']._element.rPr.rFonts.set(qn('w:eastAsia'), u'宋体')

        # 保存文件
        docx_path = user_path + name + '.docx'
        document.save(docx_path)

        i += 1


# 用户指定要处理的文件路径
file_path = 'C:\\Users\\admin\\Desktop\\学生信息表.xlsx'

# 用户指定处理完成后文件保存路径
user_path = 'C:\\Users\\admin\\Desktop\\'

# 选定用来命名输出文件的列，也是用户指定
key = '学号'

# 调用函数
tables(file_path, user_path, key)
